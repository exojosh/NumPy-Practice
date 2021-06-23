# -*- coding: utf-8 -*-
"""
Created on Tue Jun  8 16:27:39 2021

@author: Josh Limon
C/O 2021, 804-984-257

Math 142 Final Problem #2

Requries Spring data.xlsx to be in same directory as file
"""

"""
Lesson learned: Unicode encoding made this harder to code than just copying and pasting in values.
import numpy as np
import openpyxl
from pathlib import Path
import unicodedata

xlsx_file = Path('', 'Spring data.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active

x = np.zeros(10)
y1 = np.zeros(10)
y2 = np.zeros(10)
y3 = np.zeros(10)

for row in sheet.iter_rows(min_col=1,max_col=1,min_row=2,max_row=11):
    for cell in row:
        i = cell.row
        i = i -2
        x[i] = cell.value

for row in sheet.iter_rows(min_col=3,max_col=3,min_row=2,max_row=11):
    for cell in row:
        i = cell.row
        i = i -2
        y1[i] = cell.value
        
        """
import numpy as np
import openpyxl
from pathlib import Path    

x = np.linspace(-5, 4, 10)
print(x)
"""Index of 5 is left out for each value since it would be dividing by zero with current ficitonal data"""
y1 = np.zeros(10)
y1[0] = 12.981/x[0]
y1[1] = 10.4289/x[1]
y1[2] = 6.21238/x[2]
y1[3] = 7.53155/x[3]
y1[4] = 4.05626/x[4]

y1[6] = 2.19737/x[6]
y1[7] = -5.36363/x[7]
y1[8] = -7.10168/x[8]
y1[9] = -7.92041/x[9]

y2 = np.zeros(10)
y2[0] = 12.6696/x[0]
y2[1] = 9.78432/x[1]
y2[2] = 7.80185/x[2]
y2[3] = 8.04227/x[3]
y2[4] = 2.3321/x[4]

y2[6] = -2.18551/x[6]
y2[7] = -2.22239/x[7]
y2[8] = -9.45485/x[8]
y2[9] = -13.2168/x[9]

y3 = np.zeros(10)
y3[0] = 7.7284/x[0]
y3[1] = 3.28733/x[1]
y3[2] = 8.31995/x[2]
y3[3] = 1.0169/x[3]
y3[4] = 2.37081/x[4]

y3[6] = -3.45012/x[6]
y3[7] = -0.869034/x[7]
y3[8] = -7.14394/x[8]
y3[9] = -7.17117/x[9]

k1 = -sum(y1)/9
k2 = -sum(y2)/9
k3 = -sum(y3)/9
print(k1,k2,k3)
        